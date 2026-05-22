from fastapi import APIRouter, Depends, UploadFile, File, Query, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from database import get_db
from models import Product, Order, OrderItem
from datetime import date
import csv
import io
import os

router = APIRouter(prefix="/api", tags=["Import/Export"])

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "dudu_pos.db")


# ============================================================
#  EXPORT
# ============================================================

def _export_csv(headers: list[str], rows: list[list[str]], filename: str) -> StreamingResponse:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _export_xlsx(headers: list[str], rows: list[list[str]], filename: str) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="微软雅黑", size=10)
            cell.border = thin_border

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row in rows:
            val = str(row[col_idx - 1]) if row[col_idx - 1] else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _resolve_format(fmt: str, filename_csv: str, filename_xlsx: str):
    if fmt == "xlsx":
        return "xlsx", filename_xlsx
    return "csv", filename_csv


# ---- Export Products ----

@router.get("/export/products")
def export_products(fmt: str = Query("csv"), db: Session = Depends(get_db)):
    products = db.query(Product).order_by(Product.id).all()
    headers = ["ID", "品名", "规格型号", "单位", "参考单价", "创建时间", "更新时间"]
    rows = [
        [str(p.id), p.name, p.spec, p.unit, str(p.reference_price),
         str(p.created_at), str(p.updated_at)]
        for p in products
    ]
    ext, fname = _resolve_format(fmt, "products.csv", "products.xlsx")
    if ext == "xlsx":
        return _export_xlsx(headers, rows, fname)
    return _export_csv(headers, rows, fname)


# ---- Export Orders ----

def _build_receipt_xlsx(orders, filename: str) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    dark_blue = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    light_blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    white_bold = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="1F3864")
    normal_font = Font(name="微软雅黑", size=11)
    bold_font = Font(name="微软雅黑", size=11, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def fmt_yuan(v):
        return v  # keep as number for proper formatting

    col_widths = [6, 20, 14, 8, 10, 14, 16, 20]
    headers = ["行号", "品名", "规格型号", "单位", "数量", "单价（元）", "金额（元）", "备注"]
    num_cols = len(headers)

    for sheet_idx, o in enumerate(orders):
        if sheet_idx == 0:
            ws = wb.active
            ws.title = o.order_no.replace("/", "-")[:31]
        else:
            ws = wb.create_sheet(title=o.order_no.replace("/", "-")[:31])

        # Set column widths
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 1

        # Title: store name merged
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=1, value="广信区都嘟百货店")
        cell.fill = dark_blue
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
        for c in range(2, num_cols + 1):
            ws.cell(row=row, column=c).fill = dark_blue
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        # Order info row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=1, value=f"单号：{o.order_no}                    日期：{o.order_date}")
        cell.font = normal_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for c in range(2, num_cols + 1):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        # Column headers
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.fill = light_blue
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
        row += 1

        # Data rows
        for item in o.items:
            data = [
                item.row_num, item.product_name, item.spec or "", item.unit,
                item.qty, item.price, item.amount, item.remark or "",
            ]
            for i, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=i, value=val)
                cell.font = normal_font
                cell.alignment = center
                cell.border = thin_border
                if i == 5:  # qty
                    cell.number_format = '0'
                elif i in (6, 7):  # price, amount
                    cell.number_format = '#,##0.00'
            row += 1

        # Total row
        total_row_data = [
            ("总计大写", None),
            (o.amount_cn, 3),
            (o.total_qty, None),
            ("合计", None),
            (o.total_amount, None),
            ("", None),
        ]
        col = 1
        for val, merge_cols in total_row_data:
            if merge_cols and merge_cols > 1:
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + merge_cols - 1)
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = bold_font
            cell.alignment = center
            cell.border = thin_border
            if col == 5:  # qty
                cell.number_format = '0'
            elif col == 7:  # amount
                cell.number_format = '#,##0.00'
            if merge_cols and merge_cols > 1:
                for c in range(col + 1, col + merge_cols):
                    ws.cell(row=row, column=c).border = thin_border
            col += merge_cols if merge_cols else 1
        row += 1

        # Payment row
        pay_data = [
            ("收款账户", None), ("", None), ("", None),
            ("收款金额", None), ("", None),
            ("优惠金额", None), ("", None), ("", None),
        ]
        for i, (val, _) in enumerate(pay_data, 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.font = normal_font
            cell.alignment = center
            cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/orders")
def export_orders(
    fmt: str = Query("csv"),
    start: str = Query(""),
    end: str = Query(""),
    db: Session = Depends(get_db),
):
    q = db.query(Order)
    if start:
        q = q.filter(Order.order_date >= date.fromisoformat(start))
    if end:
        q = q.filter(Order.order_date <= date.fromisoformat(end))
    orders = q.order_by(Order.id).all()

    if not orders:
        raise HTTPException(status_code=404, detail="没有符合条件的单据")

    ext, fname = _resolve_format(fmt, "orders.csv", "orders.xlsx")
    if ext == "xlsx":
        return _build_receipt_xlsx(orders, fname)

    # CSV fallback — flat data export
    headers = ["单号", "日期", "总数量", "总金额", "大写金额", "行号", "品名", "规格", "单位", "数量", "单价", "金额", "备注"]
    rows = []
    for o in orders:
        for item in o.items:
            rows.append([
                o.order_no, str(o.order_date), str(o.total_qty), str(o.total_amount), o.amount_cn,
                str(item.row_num), item.product_name, item.spec, item.unit,
                str(item.qty), str(item.price), str(item.amount), item.remark,
            ])
        if not o.items:
            rows.append([
                o.order_no, str(o.order_date), str(o.total_qty), str(o.total_amount), o.amount_cn,
                "", "", "", "", "", "", "", o.remark,
            ])
    return _export_csv(headers, rows, fname)


# ---- Export Database ----

@router.get("/export/database")
def export_database():
    if not os.path.exists(DB_PATH):
        raise HTTPException(status_code=404, detail="数据库文件不存在")
    return FileResponse(
        DB_PATH,
        media_type="application/octet-stream",
        filename="dudu_pos_backup.db",
    )


# ============================================================
#  IMPORT
# ============================================================

def _parse_csv_or_xlsx(file: UploadFile) -> tuple[list[str], list[list[str]]]:
    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(c) if c is not None else "" for c in next(rows_iter)]
        data = [[str(c) if c is not None else "" for c in row] for row in rows_iter]
        return headers, data
    else:
        contents = file.file.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(contents))
        rows = list(reader)
        if not rows:
            raise HTTPException(status_code=400, detail="文件为空")
        headers = rows[0]
        data = rows[1:]
        return headers, data


def _find_col(headers: list[str], *aliases: str) -> int:
    for alias in aliases:
        for i, h in enumerate(headers):
            if h.strip() == alias:
                return i
    return -1


# ---- Import Products ----

@router.post("/import/products")
def import_products(file: UploadFile = File(...), db: Session = Depends(get_db)):
    headers, data = _parse_csv_or_xlsx(file)

    idx_name = _find_col(headers, "品名", "name")
    idx_spec = _find_col(headers, "规格型号", "规格", "spec")
    idx_unit = _find_col(headers, "单位", "unit")
    idx_price = _find_col(headers, "参考单价", "单价", "price", "reference_price")

    if idx_name < 0:
        raise HTTPException(status_code=400, detail="缺少「品名」列")

    created = 0
    for row in data:
        name = row[idx_name].strip() if idx_name < len(row) else ""
        if not name:
            continue
        spec = row[idx_spec].strip() if 0 <= idx_spec < len(row) else ""
        unit = row[idx_unit].strip() if 0 <= idx_unit < len(row) else "箱"
        try:
            price = float(row[idx_price]) if 0 <= idx_price < len(row) and row[idx_price] else 0.0
        except (ValueError, IndexError):
            price = 0.0

        # Upsert: if name exists, skip
        existing = db.query(Product).filter(Product.name == name).first()
        if existing:
            continue
        db.add(Product(name=name, spec=spec, unit=unit or "箱", reference_price=price))
        created += 1

    db.commit()
    return {"message": f"成功导入 {created} 个商品", "imported": created}


# ---- Import Orders ----

@router.post("/import/orders")
def import_orders(file: UploadFile = File(...), db: Session = Depends(get_db)):
    headers, data = _parse_csv_or_xlsx(file)

    idx_order_no = _find_col(headers, "单号", "order_no")
    idx_date = _find_col(headers, "日期", "order_date")
    idx_row_num = _find_col(headers, "行号", "row_num")
    idx_product = _find_col(headers, "品名", "product_name")
    idx_spec = _find_col(headers, "规格", "spec")
    idx_unit = _find_col(headers, "单位", "unit")
    idx_qty = _find_col(headers, "数量", "qty")
    idx_price = _find_col(headers, "单价", "price")
    idx_amount = _find_col(headers, "金额", "amount")
    idx_remark = _find_col(headers, "备注", "remark")

    if idx_order_no < 0:
        raise HTTPException(status_code=400, detail="缺少「单号」列")
    if idx_product < 0:
        raise HTTPException(status_code=400, detail="缺少「品名」列")

    # Group by order_no
    orders_map: dict[str, dict] = {}
    for row in data:
        order_no = row[idx_order_no].strip() if idx_order_no < len(row) else ""
        if not order_no:
            continue

        if order_no not in orders_map:
            orders_map[order_no] = {
                "order_no": order_no,
                "order_date": row[idx_date].strip() if 0 <= idx_date < len(row) else str(date.today()),
                "items": [],
            }

        product_name = row[idx_product].strip() if idx_product < len(row) else ""
        if not product_name:
            continue

        spec = row[idx_spec].strip() if 0 <= idx_spec < len(row) else ""
        unit = row[idx_unit].strip() if 0 <= idx_unit < len(row) else "箱"

        try:
            qty = int(float(row[idx_qty])) if 0 <= idx_qty < len(row) and row[idx_qty] else 1
        except (ValueError, IndexError):
            qty = 1
        try:
            price = float(row[idx_price]) if 0 <= idx_price < len(row) and row[idx_price] else 0.0
        except (ValueError, IndexError):
            price = 0.0
        try:
            amount = float(row[idx_amount]) if 0 <= idx_amount < len(row) and row[idx_amount] else qty * price
        except (ValueError, IndexError):
            amount = round(qty * price, 2)

        remark = row[idx_remark].strip() if 0 <= idx_remark < len(row) else ""

        orders_map[order_no]["items"].append({
            "product_name": product_name,
            "spec": spec,
            "unit": unit,
            "qty": qty,
            "price": price,
            "amount": amount,
            "remark": remark,
        })

    imported = 0
    for order_data in orders_map.values():
        existing = db.query(Order).filter(Order.order_no == order_data["order_no"]).first()
        if existing:
            continue

        total_qty = sum(it["qty"] for it in order_data["items"])
        total_amount = round(sum(it["amount"] for it in order_data["items"]), 2)

        try:
            order_date = date.fromisoformat(order_data["order_date"])
        except ValueError:
            order_date = date.today()

        order = Order(
            order_no=order_data["order_no"],
            order_date=order_date,
            total_qty=total_qty,
            total_amount=total_amount,
            amount_cn="",
            remark="",
        )
        db.add(order)
        db.flush()

        for idx, item_data in enumerate(order_data["items"], 1):
            db.add(OrderItem(
                order_id=order.id,
                row_num=idx,
                product_name=item_data["product_name"],
                spec=item_data["spec"],
                unit=item_data["unit"],
                qty=item_data["qty"],
                price=item_data["price"],
                amount=item_data["amount"],
                remark=item_data["remark"],
            ))

        imported += 1

    db.commit()
    return {"message": f"成功导入 {imported} 张单据", "imported": imported}
